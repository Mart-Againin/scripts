"""
report.py — генератор Excel-отчётов.

Структура месячного файла (monthly_YYYY-MM.xlsx):
  Лист 1: Сводка        — строка = канал
  Лист 2: Посты 24ч     — все каналы подряд, срезы ~24ч
  Лист 3: Посты-месяц   — все каналы подряд, накопленная статистика
  Лист 4: Пояснения

Структура недельного/суточного файла:
  Лист 1: Посты         — все каналы подряд
  Лист 2: По дням       — канал за каналом, 7 строк + итого
  Лист 3: По неделям    — только weekly, канал за каналом
  Лист 4: Пояснения

Источники данных:
  - 24ч срезы:          registry/<ch>/registry.json
  - Исторические:       registry/<ch>/historical/YYYY-MM.json
    (если нет — читаем из Telegram и кэшируем)

Кэш отчётов:
  Готовые файлы хранятся в output/. При повторном запросе
  предлагает отправить готовый или пересчитать.
"""

import argparse
import asyncio
import json
import logging
import sys
from calendar import monthrange
from collections import Counter
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from telethon import TelegramClient

from config import (
    API_ID, API_HASH, SESSION_NAME, CHANNELS,
    RECIPIENT_IDS, DEBUG_IDS, DEBUG_MODE,
    OUTPUT_DIR, REGISTRY_DIR, ARCHIVE_DIR, LOGS_DIR,
    TZ, CQI_W, get_telethon_kwargs,
)

logging.basicConfig(
    level=logging.DEBUG if DEBUG_MODE else logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOGS_DIR / "report.log", encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

DAYS_RU   = ["Понедельник","Вторник","Среда","Четверг","Пятница","Суббота","Воскресенье"]
MONTHS_RU = {1:"Январь",2:"Февраль",3:"Март",4:"Апрель",5:"Май",6:"Июнь",
             7:"Июль",8:"Август",9:"Сентябрь",10:"Октябрь",11:"Ноябрь",12:"Декабрь"}

# ── Стили ─────────────────────────────────────────────────────────────────
C = dict(dark="1F3864", mid="2E75B6", lite="BDD7EE", yellow="FFF2CC",
         green="E2EFDA", gray="F2F2F2", white="FFFFFF", red="FCE4D6",
         orange="F4B942", purple="D9E1F2", hist="FFF8E7")

def _f(bold=False, sz=10, color="000000", italic=False):
    return Font(name="Arial", bold=bold, size=sz, color=color, italic=italic)
def _fill(h): return PatternFill("solid", start_color=h)
def _align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _b():
    t = Side(style="thin")
    return Border(left=t, right=t, top=t, bottom=t)

def hdr(cell, text, bg=C["dark"], fg=C["white"], sz=10, bold=True, align="center"):
    cell.value = text
    cell.font  = _f(bold=bold, sz=sz, color=fg)
    cell.fill  = _fill(bg)
    cell.alignment = _align(h=align)
    cell.border = _b()

def dat(cell, text, bg=C["white"], bold=False, align="center",
        italic=False, color="000000"):
    cell.value = text
    cell.font  = _f(bold=bold, color=color, italic=italic)
    cell.fill  = _fill(bg)
    cell.alignment = _align(h=align)
    cell.border = _b()

def brd(ws, r1, r2, c1, c2):
    for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
        for cell in row:
            cell.border = _b()

def title_block(ws, text, sub, cols):
    last = get_column_letter(cols)
    ws.merge_cells(f"A1:{last}1")
    ws["A1"].value = text
    ws["A1"].font  = _f(bold=True, sz=13, color=C["white"])
    ws["A1"].fill  = _fill(C["dark"])
    ws["A1"].alignment = _align()
    ws.row_dimensions[1].height = 30
    ws.merge_cells(f"A2:{last}2")
    ws["A2"].value = sub
    ws["A2"].font  = _f(sz=9, color=C["white"])
    ws["A2"].fill  = _fill(C["mid"])
    ws["A2"].alignment = _align()
    ws.row_dimensions[2].height = 16
    ws.sheet_view.showGridLines = False
    return 3  # первая строка заголовков столбцов

# ── Метрики ───────────────────────────────────────────────────────────────
def safe_div(a, b, pct=False):
    if not b: return None
    return round((a / b) * (100 if pct else 1), 2)

def calc(snap: dict, subscribers: int) -> dict:
    v  = snap.get("views",     0)
    r  = snap.get("reactions", 0)
    c  = snap.get("comments",  0)
    f  = snap.get("forwards",  0)
    vt = snap.get("votes",     0)
    act= snap.get("actions",   0)
    sub= subscribers or 0
    cqi_raw = (r * CQI_W["react"] + vt * CQI_W["vote"] +
               f * CQI_W["forward"] + c * CQI_W["comment"])
    # Действия = только активные действия (без охвата)
    act_clean = r + c + f + vt
    return {
        "views": v, "reactions": r, "comments": c,
        "forwards": f, "votes": vt, "actions": act_clean,
        "err":    safe_div(act_clean, v,   pct=True),
        "er":     safe_div(r+c+f+vt, sub, pct=True),
        "vrpost": safe_div(v,        sub, pct=True),
        "vf":     safe_div(f,        v,   pct=True),
        "reply":  safe_div(c,        v,   pct=True),
        "poll":   safe_div(vt,       v,   pct=True) if vt > 0 else None,
        "rm":     safe_div(v,        sub),
        "cqi":    safe_div(cqi_raw,  v),
    }

def lavg(lst):
    vals = [x for x in lst if x is not None]
    return round(sum(vals) / len(vals), 2) if vals else None

def fv(cell, v, pct=False, flt=False):
    if v is None:
        cell.value = "—"; return
    if pct:
        cell.number_format = "0.00%"; cell.value = v / 100
    elif flt:
        cell.number_format = "0.00";  cell.value = v
    else:
        cell.number_format = "#,##0"; cell.value = v

# ── Столбцы постов ────────────────────────────────────────────────────────
POST_COLS = [
    ("Дата публ.",   "date",         11, False, False),
    ("Время",        "time",         10, False, False),
    ("Ссылка",       "url",          28, False, False),
    ("Тип контента", "content_type", 14, False, False),
    ("Охват",        "views",        12, False, False),
    ("Реакции",      "reactions",    11, False, False),
    ("Комменты",     "comments",     11, False, False),
    ("Пересылки",    "forwards",     12, False, False),
    ("Голоса",       "votes",        10, False, False),
    ("Действия",     "actions",      12, False, False),
    ("ERR (%)",      "err",          10, True,  False),
    ("ER (%)",       "er",           10, True,  False),
    ("VRpost (%)",   "vrpost",       11, True,  False),
    ("Viral F. (%)", "vf",           12, True,  False),
    ("Reply R. (%)", "reply",        12, True,  False),
    ("Poll R. (%)",  "poll",         11, True,  False),
    ("Reach Mult.",  "rm",           12, False, True),
    ("CQI",          "cqi",          10, False, True),
    ("Примечание",   "_note",        20, False, False),
]
N_COLS = len(POST_COLS)

# ── Реестр ────────────────────────────────────────────────────────────────
def load_registry(channel_username: str) -> dict:
    ch   = channel_username.lstrip("@")
    path = REGISTRY_DIR / ch / "registry.json"
    if not path.exists(): return {}
    try:    return json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        log.error(f"Ошибка реестра {path}: {e}"); return {}

def posts_for_period(registry: dict, date_from: date, date_to: date) -> list:
    result = []
    for p in registry.get("posts", {}).values():
        if not p.get("is_final") or not p.get("snapshot"): continue
        try:   d = datetime.strptime(p["date"], "%Y-%m-%d").date()
        except (KeyError, ValueError): continue
        if date_from <= d <= date_to:
            result.append(p)
    return sorted(result, key=lambda x: (x.get("date",""), x.get("time","")))

def archive_monthly(channel_username: str, ym: str):
    ch       = channel_username.lstrip("@")
    reg_path = REGISTRY_DIR / ch / "registry.json"
    if not reg_path.exists(): return
    registry = json.loads(reg_path.read_text(encoding="utf-8"))
    posts    = registry.get("posts", {})
    to_arch  = {mid: p for mid, p in posts.items()
                if p.get("is_final") and p.get("date","")[:7] == ym}
    remaining= {mid: p for mid, p in posts.items() if mid not in to_arch}
    if to_arch:
        arch_dir  = ARCHIVE_DIR / ch; arch_dir.mkdir(parents=True, exist_ok=True)
        arch_path = arch_dir / f"archive_{ym}.json"
        arch_path.write_text(
            json.dumps({"channel_id": channel_username, "month": ym, "posts": to_arch},
                       ensure_ascii=False, indent=2), encoding="utf-8")
        log.info(f"Архивировано {len(to_arch)} постов → {arch_path}")
    registry["posts"] = remaining
    reg_path.write_text(json.dumps(registry, ensure_ascii=False, indent=2), encoding="utf-8")

# ── Кэш отчётов ───────────────────────────────────────────────────────────
def get_cached_report_path(ym: str) -> Path | None:
    path = OUTPUT_DIR / "monthly" / f"monthly_{ym}.xlsx"
    return path if path.exists() else None

def get_cached_report_info(ym: str) -> str | None:
    path = get_cached_report_path(ym)
    if path is None: return None
    mtime = datetime.fromtimestamp(path.stat().st_mtime, tz=TZ)
    return mtime.strftime("%d.%m.%Y %H:%M")

# ── Получение данных с учётом исторического кэша ─────────────────────────
async def get_channel_posts(client, channel_username: str,
                             date_from: date, date_to: date,
                             force_historical: bool = False) -> tuple[list, int, bool]:
    """
    Возвращает (posts, subscribers, is_historical).

    Приоритет данных:
      1. 24ч-срезы из registry.json — если они есть, используются ВСЕГДА,
         независимо от force_historical. Это честные данные, собранные
         самим скриптом, их нельзя терять при пересчёте.
      2. Если 24ч-срезов нет — берём исторические данные.
         force_historical=True здесь означает «сбросить кэш historical/
         и перечитать из Telegram заново» (см. historical.py).
    """
    import historical as hist_mod

    registry    = load_registry(channel_username)
    subscribers = registry.get("subscribers", 0) if registry else 0
    posts_24h   = posts_for_period(registry, date_from, date_to) if registry else []

    if posts_24h:
        # 24ч-срезы есть — используем их в любом случае, это приоритетные данные
        return posts_24h, subscribers, False

    # 24ч-срезов нет — берём исторические (с принудительным пересбором если попросили)
    posts, subs = await hist_mod.get_posts_for_period(
        client, channel_username, date_from, date_to, force=force_historical)
    if subs: subscribers = subs
    return posts, subscribers, True

# ── Построение листа с постами всех каналов ──────────────────────────────
def _channel_divider(ws, row: int, ch_id: str, subs: int, is_hist: bool):
    last = get_column_letter(N_COLS)
    ws.merge_cells(f"A{row}:{last}{row}")
    hist_note = " │ 📜 исторические данные (статистика на дату запроса)" if is_hist else ""
    ws[f"A{row}"].value     = f"── {ch_id}  ({subs:,} подписчиков){hist_note}"
    ws[f"A{row}"].font      = _f(bold=True, sz=11, color=C["white"])
    ws[f"A{row}"].fill      = _fill(C["dark"])
    ws[f"A{row}"].alignment = _align(h="left")
    ws.row_dimensions[row].height = 22

def _write_post_row(ws, r: int, p: dict, subscribers: int, bg: str):
    sn      = p.get("snapshot", {})
    m       = calc(sn, subscribers)
    is_hist = p.get("is_historical", False)
    SNAP    = {"views","reactions","comments","forwards","votes","actions"}
    for i, (_, key, _, is_pct, is_flt) in enumerate(POST_COLS, 1):
        cell = ws.cell(row=r, column=i)
        cell.fill      = _fill(C["hist"] if is_hist else bg)
        cell.font      = _f(sz=10, italic=is_hist)
        cell.border    = _b()
        cell.alignment = _align(h="left" if i <= 4 else "center")
        if key == "_note":
            cell.value = "📜 ист." if is_hist else ""
        elif key in SNAP:
            fv(cell, sn.get(key), pct=is_pct, flt=is_flt)
        elif key in p:
            cell.value = p[key]
        else:
            fv(cell, m.get(key), pct=is_pct, flt=is_flt)

def _channel_totals_row(ws, r: int, posts: list, subscribers: int, ch_id: str):
    totals = {k: 0 for k in ["views","reactions","comments","forwards","votes","actions"]}
    mlists = {k: [] for k in ["err","er","vrpost","vf","reply","poll","rm","cqi"]}
    best   = None
    for p in posts:
        sn = p.get("snapshot", {})
        m  = calc(sn, subscribers)
        for k in totals: totals[k] += sn.get(k, 0)
        for k in mlists:
            if m[k] is not None: mlists[k].append(m[k])
        if best is None or (m["cqi"] or 0) > (calc(best.get("snapshot",{}), subscribers)["cqi"] or 0):
            best = p

    ws.merge_cells(f"A{r}:{get_column_letter(4)}{r}")
    ws[f"A{r}"].value     = f"Итого {ch_id}"
    ws[f"A{r}"].font      = _f(bold=True, sz=10)
    ws[f"A{r}"].fill      = _fill(C["yellow"])
    ws[f"A{r}"].alignment = _align()
    ws[f"A{r}"].border    = _b()
    for j in range(2, 5):
        ws.cell(row=r, column=j).fill   = _fill(C["yellow"])
        ws.cell(row=r, column=j).border = _b()

    SNAP = {"views","reactions","comments","forwards","votes","actions"}
    for i, (_, key, _, is_pct, is_flt) in enumerate(POST_COLS, 1):
        if i <= 4: continue
        cell = ws.cell(row=r, column=i)
        cell.fill      = _fill(C["yellow"])
        cell.font      = _f(bold=True, sz=10)
        cell.border    = _b()
        cell.alignment = _align()
        if key in SNAP:     fv(cell, totals[key])
        elif key in mlists: fv(cell, lavg(mlists[key]), pct=is_pct, flt=is_flt)
        else:               cell.value = "—"
    ws.row_dimensions[r].height = 20
    return best, totals, {k: lavg(v) for k, v in mlists.items()}

def build_multichannel_posts_sheet(ws, channels_data: list,
                                    title: str, subtitle: str) -> list:
    """
    Строит лист: все каналы подряд с разделителями.
    Возвращает список агрегатов для сводки.
    """
    hdr_row = title_block(ws, title, subtitle, N_COLS)
    for i, (label, _, width, _, _) in enumerate(POST_COLS, 1):
        hdr(ws[f"{get_column_letter(i)}{hdr_row}"], label, bg=C["mid"])
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[hdr_row].height = 30

    cur      = hdr_row + 1
    results  = []
    last_col = get_column_letter(N_COLS)

    for cd in channels_data:
        ch_id   = cd["channel_id"]
        subs    = cd["subscribers"]
        posts   = cd["posts"]
        is_hist = cd.get("is_historical", False)

        _channel_divider(ws, cur, ch_id, subs, is_hist); cur += 1

        if not posts:
            ws.merge_cells(f"A{cur}:{last_col}{cur}")
            ws[f"A{cur}"].value     = "Нет данных за период"
            ws[f"A{cur}"].font      = _f(sz=10, italic=True, color="888888")
            ws[f"A{cur}"].fill      = _fill(C["gray"])
            ws[f"A{cur}"].alignment = _align(h="left")
            ws.row_dimensions[cur].height = 18
            cur += 2
            continue

        for idx, p in enumerate(posts):
            _write_post_row(ws, cur, p, subs, C["white"] if idx%2==0 else C["gray"])
            ws.row_dimensions[cur].height = 18
            cur += 1

        best, totals, avgs = _channel_totals_row(ws, cur, posts, subs, ch_id)
        cur += 1

        if best:
            ws.merge_cells(f"A{cur}:{last_col}{cur}")
            ws[f"A{cur}"].value = (f"🏆 Лучший по CQI: {best.get('url','')} "
                f"({best.get('date','')} {best.get('time','')} | {best.get('content_type','')})")
            ws[f"A{cur}"].font      = _f(sz=9, italic=True, color=C["dark"])
            ws[f"A{cur}"].fill      = _fill(C["green"])
            ws[f"A{cur}"].alignment = _align(h="left")
            ws.row_dimensions[cur].height = 16
            cur += 1

        cur += 1  # пустая строка между каналами
        results.append({"channel_id": ch_id, "subscribers": subs,
                         "posts": posts, "is_historical": is_hist,
                         "totals": totals, "avgs": avgs, "best": best,
                         "count": len(posts)})

    if any(r.get("is_historical") for r in results):
        ws.merge_cells(f"A{cur}:{last_col}{cur}")
        ws[f"A{cur}"].value = ("📜 Жёлтый фон = исторические данные. "
            "Статистика на момент запроса, а не через 24 ч после публикации.")
        ws[f"A{cur}"].font      = _f(sz=9, italic=True, color="7D6608")
        ws[f"A{cur}"].fill      = _fill(C["hist"])
        ws[f"A{cur}"].alignment = _align(h="left")
        ws.row_dimensions[cur].height = 18

    return results

# ── Сводный лист ──────────────────────────────────────────────────────────
SUMMARY_COLS = [
    ("Канал",              22), ("Подписчики",      13), ("+месяц",           10),
    ("Постов",             10), ("Сторис",           10),
    ("Охват постов",       14), ("Охват сторис",     14), ("Общий охват",      14),
    ("Реакции (сумм.)",    13), ("Комменты (сумм.)", 13),
    ("Пересылки (сумм.)",  14), ("Голоса (сумм.)",   12), ("Действия (сумм.)", 14),
    ("Ср. ERR (%)",        11), ("Ср. ER (%)",       11),
    ("Ср. VRpost (%)",     12), ("Ср. Viral F.(%)",  13), ("Ср. Reply R.(%)",  13),
    ("Ср. Reach Mult.",    13), ("Ср. CQI",          10), ("Топ формат",       14),
    ("Лучший пост (CQI)",  28), ("Данные",           16),
]
# Индексы под новую раскладку (считая с 1)
PCT_SUM   = {14,15,16,17,18}
FLOAT_SUM = {19,20}

def build_summary_sheet(ws, results: list, label: str, subtitle: str):
    n = len(SUMMARY_COLS)
    hdr_row = title_block(ws, f"📊 СВОДКА — {label}", subtitle, n)
    for i, (lbl, width) in enumerate(SUMMARY_COLS, 1):
        hdr(ws[f"{get_column_letter(i)}{hdr_row}"], lbl, bg=C["mid"])
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[hdr_row].height = 30

    GROWTH_COL = 3  # +месяц — может быть отрицательным

    for idx, cr in enumerate(results):
        r       = hdr_row + 1 + idx
        bg      = C["white"] if idx%2==0 else C["gray"]
        t       = cr["totals"]; a = cr["avgs"]
        tf      = Counter(p.get("content_type","") for p in cr["posts"]).most_common(1)
        top_fmt = tf[0][0] if tf else "—"
        best_url= cr["best"]["url"] if cr.get("best") else "—"
        dtype   = "📜 Исторические" if cr.get("is_historical") else "✅ 24ч срезы"

        growth       = cr.get("growth", {})
        stories_data = cr.get("stories", {})
        views_posts  = t.get("views") or 0
        views_stories= stories_data.get("views", 0)
        views_total  = views_posts + views_stories

        row_v = [
            cr["channel_id"], cr["subscribers"], growth.get("month"),
            cr["count"], stories_data.get("count", 0),
            views_posts, views_stories, views_total,
            t.get("reactions"), t.get("comments"),
            t.get("forwards"), t.get("votes"), t.get("actions"),
            a.get("err"), a.get("er"), a.get("vrpost"),
            a.get("vf"), a.get("reply"), a.get("rm"), a.get("cqi"),
            top_fmt, best_url, dtype,
        ]
        for i, v in enumerate(row_v, 1):
            cell = ws.cell(row=r, column=i)
            cell.fill      = _fill(bg)
            cell.font      = _f(sz=10)
            cell.border    = _b()
            cell.alignment = _align(h="left" if i==1 else "center")
            if v is None:
                cell.value = "—"
            elif i == GROWTH_COL:
                cell.value = f"+{v}" if v > 0 else str(v)
                cell.font  = _f(sz=10, bold=True,
                                color="2E7D32" if v > 0 else ("C62828" if v < 0 else "666666"))
            elif i in PCT_SUM:
                fv(cell, v, pct=True)
            elif i in FLOAT_SUM:
                fv(cell, v, flt=True)
            elif i in {2,4,5,6,7,8,9,10,11,12,13}:
                fv(cell, v)
            else:
                cell.value = v
        ws.row_dimensions[r].height = 22

    rt = hdr_row + 1 + len(results)
    ws.merge_cells(f"A{rt}:B{rt}")
    ws[f"A{rt}"].value     = "ИТОГО / СРЕДНЕЕ"
    ws[f"A{rt}"].font      = _f(bold=True, sz=10)
    ws[f"A{rt}"].fill      = _fill(C["yellow"])
    ws[f"A{rt}"].alignment = _align()
    ws[f"A{rt}"].border    = _b()

    keys = ["","","","count","_stories_count","_views_posts","_views_stories","_views_total",
            "reactions","comments","forwards","votes","actions",
            "err","er","vrpost","vf","reply","rm","cqi"]
    for j in range(3, n+1):
        cell = ws.cell(row=rt, column=j)
        cell.fill = _fill(C["yellow"]); cell.font = _f(bold=True, sz=10)
        cell.border = _b(); cell.alignment = _align()
        key = keys[j-1] if j-1 < len(keys) else ""

        if j == GROWTH_COL:
            cell.value = "—"
        elif key == "count":
            fv(cell, sum(cr["count"] for cr in results))
        elif key == "_stories_count":
            fv(cell, sum(cr.get("stories",{}).get("count",0) for cr in results))
        elif key == "_views_posts":
            fv(cell, sum(cr["totals"].get("views",0) for cr in results))
        elif key == "_views_stories":
            fv(cell, sum(cr.get("stories",{}).get("views",0) for cr in results))
        elif key == "_views_total":
            tv = sum(cr["totals"].get("views",0) for cr in results)
            sv = sum(cr.get("stories",{}).get("views",0) for cr in results)
            fv(cell, tv+sv)
        elif key in {"reactions","comments","forwards","votes","actions"}:
            fv(cell, sum(cr["totals"].get(key,0) for cr in results))
        elif j in PCT_SUM:
            fv(cell, lavg([cr["avgs"].get(key) for cr in results]), pct=True)
        elif j in FLOAT_SUM:
            fv(cell, lavg([cr["avgs"].get(key) for cr in results]), flt=True)
        else:
            cell.value = "—"
    ws.row_dimensions[rt].height = 22
    brd(ws, hdr_row, rt, 1, n)

    # Пометка о динамике подписчиков
    rn = rt + 2
    ws.merge_cells(f"A{rn}:{get_column_letter(n)}{rn}")
    ws[f"A{rn}"].value = ("ℹ️  Колонка «+месяц» показывает прирост подписчиков с прошлой месячной записи. "
        "Данные накапливаются с момента запуска скрипта — если истории меньше нужного периода, ячейка будет «—».")
    ws[f"A{rn}"].font      = _f(sz=9, italic=True, color="595959")
    ws[f"A{rn}"].alignment = _align(h="left")
    ws.row_dimensions[rn].height = 18

# ── Агрегатный лист (По дням / По неделям) — канал за каналом ────────────
AGG_COLS_DAY = [
    ("Дата",          10), ("День недели",    14), ("Постов",  8),
    ("Охват",         12), ("Ср. охват",      12), ("Реакции", 11),
    ("Комменты",      11), ("Пересылки",      12), ("Голоса",  10),
    ("Действия",      12), ("Ср. ERR (%)",    11), ("Ср. ER (%)",11),
    ("Ср. VRpost (%)",12), ("Ср. Viral (%)",  12), ("Ср. CQI", 10),
    ("Лучший пост",   28),
]
AGG_COLS_WEEK = [
    ("Неделя",        20), ("Постов",   8),
    ("Охват",         12), ("Реакции",  11), ("Комменты", 11),
    ("Пересылки",     12), ("Голоса",   10), ("Действия", 12),
    ("Ср. ERR (%)",   11), ("Ср. ER (%)",11),("Ср. VRpost (%)",12),
    ("Ср. CQI",       10), ("Лучший пост", 28),
]

def _agg_bucket_row(ws, r: int, label: str, dow: str, posts: list,
                     subscribers: int, bg: str, mode: str):
    """mode = 'day' или 'week'"""
    cols = AGG_COLS_DAY if mode=="day" else AGG_COLS_WEEK
    sn_all  = [p.get("snapshot",{}) for p in posts]
    metrics = [calc(sn, subscribers) for sn in sn_all]
    views_s = sum(sn.get("views",0) for sn in sn_all)
    best    = max(posts, key=lambda p: calc(p.get("snapshot",{}),subscribers)["cqi"] or 0) if posts else None

    def _cv(col_idx, v, pct=False, flt=False):
        cell = ws.cell(row=r, column=col_idx)
        cell.fill = _fill(bg); cell.font = _f(sz=10)
        cell.border = _b(); cell.alignment = _align(h="center")
        if v is None or (isinstance(v,int) and v==0):
            cell.value = "—" if col_idx > 2 else v
        elif pct: fv(cell, v, pct=True)
        elif flt: fv(cell, v, flt=True)
        else: cell.value = v

    if mode == "day":
        _cv(1, label, False, False); ws.cell(row=r,column=1).alignment=_align(h="left")
        _cv(2, dow,   False, False)
        _cv(3, len(posts))
        _cv(4, views_s or None)
        _cv(5, round(views_s/len(posts),1) if posts else None, flt=True)
        _cv(6, sum(sn.get("reactions",0) for sn in sn_all) or None)
        _cv(7, sum(sn.get("comments",0)  for sn in sn_all) or None)
        _cv(8, sum(sn.get("forwards",0)  for sn in sn_all) or None)
        _cv(9, sum(sn.get("votes",0)     for sn in sn_all) or None)
        _cv(10,sum(sn.get("actions",0)   for sn in sn_all) or None)
        _cv(11,lavg([m["err"]    for m in metrics]), pct=True)
        _cv(12,lavg([m["er"]     for m in metrics]), pct=True)
        _cv(13,lavg([m["vrpost"] for m in metrics]), pct=True)
        _cv(14,lavg([m["vf"]     for m in metrics]), pct=True)
        _cv(15,lavg([m["cqi"]    for m in metrics]), flt=True)
        ws.cell(row=r,column=16).value  = best["url"] if best else "—"
        ws.cell(row=r,column=16).fill   = _fill(bg)
        ws.cell(row=r,column=16).font   = _f(sz=10)
        ws.cell(row=r,column=16).border = _b()
        ws.cell(row=r,column=16).alignment = _align(h="left")
    else:
        _cv(1, label, False, False); ws.cell(row=r,column=1).alignment=_align(h="left")
        _cv(2, len(posts))
        _cv(3, views_s or None)
        _cv(4, sum(sn.get("reactions",0) for sn in sn_all) or None)
        _cv(5, sum(sn.get("comments",0)  for sn in sn_all) or None)
        _cv(6, sum(sn.get("forwards",0)  for sn in sn_all) or None)
        _cv(7, sum(sn.get("votes",0)     for sn in sn_all) or None)
        _cv(8, sum(sn.get("actions",0)   for sn in sn_all) or None)
        _cv(9, lavg([m["err"]    for m in metrics]), pct=True)
        _cv(10,lavg([m["er"]     for m in metrics]), pct=True)
        _cv(11,lavg([m["vrpost"] for m in metrics]), pct=True)
        _cv(12,lavg([m["cqi"]    for m in metrics]), flt=True)
        ws.cell(row=r,column=13).value  = best["url"] if best else "—"
        ws.cell(row=r,column=13).fill   = _fill(bg)
        ws.cell(row=r,column=13).font   = _f(sz=10)
        ws.cell(row=r,column=13).border = _b()
        ws.cell(row=r,column=13).alignment = _align(h="left")

    ws.row_dimensions[r].height = 18

def build_agg_multichannel(ws, channels_data: list,
                            date_from: date, date_to: date,
                            title: str, subtitle: str, mode: str):
    """
    mode = 'day' или 'week'
    Строит агрегатный лист: канал за каналом с разделителями.
    """
    cols = AGG_COLS_DAY if mode=="day" else AGG_COLS_WEEK
    n    = len(cols)
    hdr_row = title_block(ws, title, subtitle, n)
    for i,(label,width) in enumerate(cols,1):
        hdr(ws[f"{get_column_letter(i)}{hdr_row}"], label, bg=C["mid"])
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[hdr_row].height = 30

    last_col = get_column_letter(n)
    cur      = hdr_row + 1

    for cd in channels_data:
        ch_id   = cd["channel_id"]
        subs    = cd["subscribers"]
        posts   = cd["posts"]
        is_hist = cd.get("is_historical", False)

        _channel_divider(ws, cur, ch_id, subs, is_hist); cur += 1

        if not posts:
            ws.merge_cells(f"A{cur}:{last_col}{cur}")
            ws[f"A{cur}"].value     = "Нет данных за период"
            ws[f"A{cur}"].font      = _f(sz=10, italic=True, color="888888")
            ws[f"A{cur}"].fill      = _fill(C["gray"])
            ws[f"A{cur}"].alignment = _align(h="left")
            ws.row_dimensions[cur].height = 18
            cur += 2; continue

        # Группируем посты по нужному периоду
        by_date = {}
        for p in posts:
            by_date.setdefault(p.get("date",""), []).append(p)

        data_start_row = cur  # первая строка с данными — нужна для графика

        if mode == "day":
            # Итерируем по дням периода
            d = date_from
            idx = 0
            while d <= date_to:
                ds   = d.strftime("%Y-%m-%d")
                dow  = d.weekday()
                dp   = by_date.get(ds, [])
                is_we= dow >= 5
                bg   = C["red"] if is_we else (C["white"] if idx%2==0 else C["gray"])
                _agg_bucket_row(ws, cur, d.strftime("%d.%m.%Y"),
                                DAYS_RU[dow], dp, subs, bg, "day")
                cur += 1; d += timedelta(days=1); idx += 1
        else:
            # Недели
            week_start = date_from - timedelta(days=date_from.weekday())
            idx = 0
            while week_start <= date_to:
                wend  = min(week_start + timedelta(days=6), date_to)
                wp    = []
                d     = week_start
                while d <= wend:
                    wp.extend(by_date.get(d.strftime("%Y-%m-%d"), []))
                    d += timedelta(days=1)
                label = f"{week_start.strftime('%d.%m')} – {wend.strftime('%d.%m.%Y')}"
                bg    = C["white"] if idx%2==0 else C["gray"]
                _agg_bucket_row(ws, cur, label, "", wp, subs, bg, "week")
                cur += 1; week_start += timedelta(days=7); idx += 1

        data_end_row = cur - 1  # последняя строка с данными (перед "Итого")

        # Итого по каналу
        ws.merge_cells(f"A{cur}:{last_col}{cur}")
        ws[f"A{cur}"].value     = f"Итого {ch_id}"
        ws[f"A{cur}"].font      = _f(bold=True, sz=10)
        ws[f"A{cur}"].fill      = _fill(C["yellow"])
        ws[f"A{cur}"].alignment = _align(h="left")
        ws[f"A{cur}"].border    = _b()
        for j in range(2, n+1):
            ws.cell(row=cur, column=j).fill   = _fill(C["yellow"])
            ws.cell(row=cur, column=j).border = _b()
        ws.row_dimensions[cur].height = 20
        cur += 2  # итого + пустая строка

        # Графики по каналу: охват по периодам + ERR по периодам
        if data_end_row >= data_start_row:
            views_col = 4 if mode == "day" else 3
            err_col   = 11 if mode == "day" else 9
            cat_col   = 1  # столбец с датой/неделей — категории для оси X

            cats_ref = Reference(ws, min_col=cat_col, min_row=data_start_row, max_row=data_end_row)

            chart_views = BarChart()
            chart_views.type = "col"
            chart_views.title = f"{ch_id} — охват по периодам"
            chart_views.y_axis.title = "Просмотры"
            chart_views.height = 7
            chart_views.width = 16
            views_ref = Reference(ws, min_col=views_col, min_row=hdr_row, max_row=data_end_row)
            chart_views.add_data(views_ref, titles_from_data=True)
            chart_views.set_categories(cats_ref)
            if chart_views.series:
                chart_views.series[0].graphicalProperties.solidFill = C["mid"]
            chart_views.legend = None
            ws.add_chart(chart_views, f"A{cur}")

            chart_err = LineChart()
            chart_err.title = f"{ch_id} — ERR (%) по периодам"
            chart_err.y_axis.title = "ERR %"
            chart_err.height = 7
            chart_err.width = 16
            err_ref = Reference(ws, min_col=err_col, min_row=hdr_row, max_row=data_end_row)
            chart_err.add_data(err_ref, titles_from_data=True)
            chart_err.set_categories(cats_ref)
            if chart_err.series:
                chart_err.series[0].graphicalProperties.line.solidFill = C["orange"]
            chart_err.legend = None
            # Графики ставим бок о бок (примерно 8 колонок ширина бар-чарта в Excel-единицах)
            anchor_col = get_column_letter(n + 2)
            ws.add_chart(chart_err, f"{anchor_col}{cur}")

            # Резервируем строки под высоту графиков (примерно 14-15 строк на каждый)
            cur += 16

    # Заметка о выходных (только для дневного)
    if mode == "day":
        ws.merge_cells(f"A{cur}:{last_col}{cur}")
        ws[f"A{cur}"].value     = "🔴 Красным выделены выходные дни"
        ws[f"A{cur}"].font      = _f(sz=9, italic=True, color="C00000")
        ws[f"A{cur}"].alignment = _align(h="left")
        ws.row_dimensions[cur].height = 14


# ── Лист «Календарь» (только для месячного отчёта) ────────────────────────

WEEKDAY_LABELS = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]

def build_calendar_sheet(ws, channels_data: list,
                          date_from: date, date_to: date,
                          title: str, subtitle: str):
    """
    Календарная сетка по месяцу: строки = недели (1-5), столбцы = дни недели (Пн-Вс).
    Один календарь на канал, каналы идут друг за другом сверху вниз.
    В ячейке дня — количество постов или «—» если дня нет в периоде/постов не было.
    Дни вне месяца (до 1-го числа или после последнего) помечаются отдельным цветом.
    """
    n = 7  # 7 дней недели
    hdr_row = title_block(ws, title, subtitle, n)

    for i, label in enumerate(WEEKDAY_LABELS, 1):
        hdr(ws[f"{get_column_letter(i)}{hdr_row}"], label, bg=C["mid"])
        ws.column_dimensions[get_column_letter(i)].width = 13
    ws.row_dimensions[hdr_row].height = 24

    # Группируем посты по дате для быстрого доступа
    cur = hdr_row + 1
    last_col = get_column_letter(n)

    # Первая и последняя неделя месяца (пн-вс), чтобы знать сколько строк-недель нужно
    grid_start = date_from - timedelta(days=date_from.weekday())          # понедельник первой недели
    grid_end   = date_to + timedelta(days=(6 - date_to.weekday()))        # воскресенье последней недели
    n_weeks = (grid_end - grid_start).days // 7 + 1

    for cd in channels_data:
        ch_id   = cd["channel_id"]
        subs    = cd["subscribers"]
        posts   = cd["posts"]
        is_hist = cd.get("is_historical", False)

        by_date = {}
        for p in posts:
            by_date.setdefault(p.get("date", ""), []).append(p)

        # Заголовок канала
        _channel_divider(ws, cur, ch_id, subs, is_hist)
        cur += 1

        # Сетка недель
        week_start = grid_start
        for w in range(n_weeks):
            for d_idx in range(7):
                day = week_start + timedelta(days=d_idx)
                col = d_idx + 1
                cell_ref = ws.cell(row=cur, column=col)
                cell_ref.border = _b()
                cell_ref.alignment = _align()

                in_month = date_from <= day <= date_to
                if not in_month:
                    # День вне диапазона месяца — серая ячейка с прочерком
                    cell_ref.value = "—"
                    cell_ref.fill = _fill("EDEDED")
                    cell_ref.font = _f(sz=9, color="B0B0B0", italic=True)
                else:
                    ds = day.strftime("%Y-%m-%d")
                    day_posts = by_date.get(ds, [])
                    is_weekend = d_idx >= 5
                    count = len(day_posts)

                    if count == 0:
                        cell_ref.value = f"{day.day}\n—"
                        cell_ref.fill = _fill(C["red"] if is_weekend else C["white"])
                        cell_ref.font = _f(sz=10, color="999999")
                    else:
                        cell_ref.value = f"{day.day}\n{count} пост." if count != 1 else f"{day.day}\n1 пост"
                        cell_ref.fill = _fill(C["green"] if not is_weekend else C["red"])
                        cell_ref.font = _f(sz=10, bold=True)
            ws.row_dimensions[cur].height = 32
            cur += 1
            week_start += timedelta(days=7)

        # Лёгкая разделительная строка между календарями каналов
        cur += 1

    # Легенда
    ws.merge_cells(f"A{cur}:{last_col}{cur}")
    ws[f"A{cur}"].value = ("Серым — дни вне месяца. Зелёным — дни с постами. "
                            "Белым/красным (выходные) — дни без публикаций.")
    ws[f"A{cur}"].font = _f(sz=9, italic=True, color="595959")
    ws[f"A{cur}"].alignment = _align(h="left")
    ws.row_dimensions[cur].height = 18
    ws.sheet_view.showGridLines = False

# ── Лист пояснений ────────────────────────────────────────────────────────
LEGEND = [
    ("+месяц", "Подписчики этот месяц − подписчики прошлый месяц",
     "Динамика подписчиков. Фиксируется раз в месяц (в начале месяца). Накапливается с момента запуска скрипта — если данных меньше двух месяцев, будет «—»", "Базовые данные"),
    ("Сторис",                "Telegram API: stories.GetPinnedStories / GetStoriesArchive",
     "Количество сторис опубликованных каналом за период. Учитываются только сторис, отслеженные скриптом (с момента запуска)", "Базовые данные"),
    ("Охват постов",          "Сумма Views по всем постам периода",
     "Просмотры обычных постов канала", "Базовые данные"),
    ("Охват сторис",          "Сумма просмотров сторис за период",
     "Доступно для всех публичных сторис. Реакции на сторис доступны только если вы администратор канала", "Базовые данные"),
    ("Общий охват",           "Охват постов + Охват сторис",
     "Суммарный охват всего контента канала за период", "Базовые данные"),
    ("Охват (Views)",        "Telegram API: message.views",
     "Суммарное число просмотров поста на момент сбора данных", "Базовые данные"),
    ("Реакции (React)",      "Telegram API: сумма всех реакций",
     "Все эмодзи-реакции под постом: 👍 ❤️ 🔥 и т.д.", "Базовые данные"),
    ("Комментарии",          "Telegram API: replies.replies",
     "Число комментариев в linked-группе канала", "Базовые данные"),
    ("Пересылки (Forwards)", "Telegram API: message.forwards",
     "Сколько раз пост был переслан", "Базовые данные"),
    ("Голоса (Votes)",       "Telegram API: сумма voters по вариантам опроса",
     "Сумма всех голосов в опросе. Для постов без опроса = 0", "Базовые данные"),
    ("Число действий",       "Views + React + Comments + Forwards + Votes",
     "Совокупная активность вокруг поста", "Базовые данные"),
    ("ERR (%)",              "Actions / Views × 100%",
     "Доля просмотревших, совершивших хоть какое-то действие. Норма: 3–8%", "Коэффициенты"),
    ("ER (%)",               "(React+Comments+Fwd+Votes) / Подписчики × 100%",
     "Классический ER. Не зависит от алгоритмов показа", "Коэффициенты"),
    ("VRpost (%)",           "Views / Подписчики × 100%",
     "Какой % подписчиков увидел пост", "Коэффициенты"),
    ("Viral Factor (%)",     "Forwards / Views × 100%",
     "Доля зрителей, поделившихся постом", "Виральность"),
    ("Reply Rate (%)",       "Comments / Views × 100%",
     "Насколько пост провоцирует диалог", "Виральность"),
    ("Poll Rate (%)",        "Votes / Views × 100%",
     "Только для опросов. Доля проголосовавших среди просмотревших", "Виральность"),
    ("Reach Multiplier",     "Views / Подписчики",
     "Коэффициент > 1: пост вышел за пределы базы подписчиков", "Виральность"),
    ("CQI",
     f"(React×{CQI_W['react']}+Votes×{CQI_W['vote']}+Fwd×{CQI_W['forward']}+Comments×{CQI_W['comment']}) / Views × 100",
     "Взвешенный индекс качества. Веса настраиваются в .env", "Индексы"),
    ("📜 Исторические данные", "Статистика на дату запроса отчёта",
     "Посты собраны из истории канала. Статистика накопленная, а не суточная. "
     "Отмечены жёлтым фоном и пометкой «ист.»", "Служебные"),
]

BLOCK_C = {"Базовые данные":C["lite"],"Коэффициенты":C["purple"],
           "Виральность":C["green"],"Индексы":C["yellow"],"Служебные":C["gray"]}

def build_legend_sheet(wb: Workbook):
    ws = wb.create_sheet("Пояснения к показателям")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:D1")
    ws["A1"].value     = "📖 ПОЯСНЕНИЯ К ПОКАЗАТЕЛЯМ И ФОРМУЛАМ"
    ws["A1"].font      = _f(bold=True, sz=13, color=C["white"])
    ws["A1"].fill      = _fill(C["dark"])
    ws["A1"].alignment = _align()
    ws.row_dimensions[1].height = 30

    for col_l, text, width in [("A","Показатель",26),("B","Формула / Источник",40),
                                 ("C","Описание",62),("D","Блок",20)]:
        hdr(ws[f"{col_l}2"], text, bg=C["mid"])
        ws.column_dimensions[col_l].width = width
    ws.row_dimensions[2].height = 20

    for i, (name, formula, desc, block) in enumerate(LEGEND):
        r  = 3 + i
        bg = C["white"] if i%2==0 else C["gray"]
        dat(ws[f"A{r}"], name,    bg=bg, bold=True,   align="left")
        dat(ws[f"B{r}"], formula, bg=bg, italic=True,  color=C["dark"], align="left")
        dat(ws[f"C{r}"], desc,    bg=bg, align="left")
        ws[f"D{r}"].value     = block
        ws[f"D{r}"].font      = _f(sz=9, bold=True)
        ws[f"D{r}"].fill      = _fill(BLOCK_C.get(block, C["white"]))
        ws[f"D{r}"].alignment = _align()
        ws[f"D{r}"].border    = _b()
        ws.row_dimensions[r].height = 36
    brd(ws, 2, 2+len(LEGEND), 1, 4)

    CQI_W_TABLE = [
        ("Реакция (лайк, эмодзи)",  f"× {CQI_W['react']}",   "Минимальное усилие"),
        ("Голос в опросе",           f"× {CQI_W['vote']}",    "Требует прочтения вариантов"),
        ("Пересылка (Forward)",      f"× {CQI_W['forward']}", "Решение поделиться = высокая оценка"),
        ("Комментарий",              f"× {CQI_W['comment']}", "Максимальное усилие — написать текст"),
    ]
    rc = 3 + len(LEGEND) + 2
    ws.merge_cells(f"A{rc}:D{rc}")
    ws[f"A{rc}"].value     = "⚖️  ВЕСА CQI — настраиваются в .env"
    ws[f"A{rc}"].font      = _f(bold=True, sz=10, color=C["white"])
    ws[f"A{rc}"].fill      = _fill(C["dark"])
    ws[f"A{rc}"].alignment = _align(h="left")
    ws.row_dimensions[rc].height = 22
    for j, (act, w, reason) in enumerate(CQI_W_TABLE):
        r2 = rc + 1 + j
        bg = C["white"] if j%2==0 else C["gray"]
        dat(ws[f"A{r2}"], act,    bg=bg, align="left")
        dat(ws[f"B{r2}"], w,      bg=bg, bold=True)
        dat(ws[f"C{r2}"], reason, bg=bg, align="left")
        ws[f"D{r2}"].fill   = _fill(bg)
        ws[f"D{r2}"].border = _b()
        ws.row_dimensions[r2].height = 20
    brd(ws, rc, rc+len(CQI_W_TABLE), 1, 4)

# ── Определение периода ───────────────────────────────────────────────────
def determine_period(report_type: str,
                     week_offset: int = 1,
                     week_date: date = None) -> tuple:
    """
    Возвращает (date_from, date_to, period_label).
    week_offset: 1=прошлая неделя, 2=позапрошлая
    week_date: любой день нужной недели (для ручного ввода)
    """
    today = date.today()
    if report_type == "daily":
        d = today - timedelta(days=1)
        return d, d, d.strftime("%d.%m.%Y")
    elif report_type == "weekly":
        if week_date:
            mon = week_date - timedelta(days=week_date.weekday())
        else:
            mon = today - timedelta(days=today.weekday() + 7 * week_offset)
        sun = mon + timedelta(days=6)
        return mon, sun, f"{mon.strftime('%d.%m')}–{sun.strftime('%d.%m.%Y')}"
    elif report_type == "monthly":
        first = today.replace(day=1) - timedelta(days=1)
        d_from = first.replace(day=1)
        return d_from, first, f"{MONTHS_RU.get(d_from.month,'')} {d_from.year}"
    raise ValueError(f"Неизвестный тип: {report_type}")

# ── Главная функция ───────────────────────────────────────────────────────
async def build_and_send(report_type: str, debug_override: bool = False,
                         month_override: str = None, tg_client=None,
                         week_offset: int = 1, week_date: date = None,
                         force_rebuild: bool = False,
                         override_recipients: list = None):
    is_debug   = DEBUG_MODE or debug_override
    # override_recipients — точечная отправка конкретному человеку
    # (например, модератор запросил суточный отчёт лично себе)
    recipients = override_recipients if override_recipients is not None else (
        DEBUG_IDS if is_debug else RECIPIENT_IDS)
    channels   = CHANNELS
    collect_ts = datetime.now(TZ).strftime("%d.%m.%Y %H:%M")

    if month_override and report_type == "monthly":
        year, mo = int(month_override[:4]), int(month_override[5:7])
        d_from   = date(year, mo, 1)
        d_to     = date(year, mo, monthrange(year, mo)[1])
        period_label = f"{MONTHS_RU.get(mo,'')} {year}"
    else:
        d_from, d_to, period_label = determine_period(
            report_type, week_offset=week_offset, week_date=week_date)

    ym         = d_from.strftime("%Y-%m")
    month_name = MONTHS_RU.get(d_from.month, str(d_from.month))

    log.info(f"=== Отчёт {report_type} | {period_label} | debug={is_debug} ===")

    out_dir = OUTPUT_DIR / {"daily":"daily","weekly":"weekly","monthly":"monthly"}[report_type]
    out_dir.mkdir(parents=True, exist_ok=True)

    async def _execute(client):
        channels_data = []
        for ch in channels:
            posts, subs, is_hist = await get_channel_posts(
                client, ch, d_from, d_to, force_historical=force_rebuild)
            channels_data.append({
                "channel_id":   ch,
                "subscribers":  subs,
                "posts":        posts,
                "is_historical": is_hist,
            })

        wb = Workbook(); wb.remove(wb.active)

        if report_type == "monthly":
            # Лист 1: Сводка
            ws_sum = wb.create_sheet("Сводка")
            results_24h = build_multichannel_posts_sheet(
                wb.create_sheet("Посты 24ч"), channels_data,
                f"📋 ПОСТЫ 24Ч — {month_name.upper()} {d_from.year}",
                "Статистика через ~24 ч после публикации. Жёлтый = исторические данные.",
            )
            results_mo = build_multichannel_posts_sheet(
                wb.create_sheet("Посты-месяц"), channels_data,
                f"📋 ПОСТЫ-МЕСЯЦ — {month_name.upper()} {d_from.year}",
                "Накопленная статистика на дату сбора.",
            )
            # Сводку строим из results_mo
            # Подмешиваем данные о приросте подписчиков и сторис
            import snapshot as _snap_mod

            try:
                import stories as _stories_mod
            except Exception as e:
                log.warning(f"Модуль stories.py недоступен: {e}")
                _stories_mod = None

            for cr in results_mo:
                ch_id = cr["channel_id"]
                cr["growth"] = _snap_mod.get_subscriber_growth(ch_id)
                if _stories_mod:
                    try:
                        cr["stories"] = _stories_mod.get_stories_summary(ch_id, d_from, d_to)
                    except Exception as e:
                        log.warning(f"Сторис {ch_id}: {e}")
                        cr["stories"] = {"count": 0, "views": 0, "reactions": 0}
                else:
                    cr["stories"] = {"count": 0, "views": 0, "reactions": 0}

            build_summary_sheet(
                ws_sum, results_mo,
                f"{month_name.upper()} {d_from.year}",
                f"Период: {d_from.strftime('%d.%m.%Y')} – {d_to.strftime('%d.%m.%Y')} | Собрано: {collect_ts}",
            )
            # Листы По дням и По неделям
            build_agg_multichannel(
                wb.create_sheet("По дням"), channels_data,
                d_from, d_to,
                f"📅 ПО ДНЯМ — {month_name.upper()} {d_from.year}",
                "Каждая строка = один день. Канал за каналом. 🔴 Красным — выходные.",
                mode="day",
            )
            build_agg_multichannel(
                wb.create_sheet("По неделям"), channels_data,
                d_from, d_to,
                f"📆 ПО НЕДЕЛЯМ — {month_name.upper()} {d_from.year}",
                "Каждая строка = одна неделя. Канал за каналом.",
                mode="week",
            )
            # Лист «Календарь» — наглядная сетка по дням месяца, канал за каналом
            build_calendar_sheet(
                wb.create_sheet("Календарь"), channels_data,
                d_from, d_to,
                f"🗓️ КАЛЕНДАРЬ ПУБЛИКАЦИЙ — {month_name.upper()} {d_from.year}",
                "Недели по вертикали, дни недели по горизонтали. Число — количество постов за день.",
            )
            # Переставляем листы: Сводка первой
            wb.move_sheet("Сводка", offset=-len(wb.sheetnames)+1)

        else:
            # daily / weekly
            build_multichannel_posts_sheet(
                wb.create_sheet("Посты"), channels_data,
                f"📋 ПОСТЫ — {period_label}",
                "Статистика через ~24 ч после публикации.",
            )
            build_agg_multichannel(
                wb.create_sheet("По дням"), channels_data,
                d_from, d_to,
                f"📅 ПО ДНЯМ — {period_label}",
                "Каждая строка = один день. Канал за каналом.",
                mode="day",
            )
            if report_type == "weekly":
                build_agg_multichannel(
                    wb.create_sheet("По неделям"), channels_data,
                    d_from, d_to,
                    f"📆 ПО НЕДЕЛЯМ — {period_label}",
                    "Каждая строка = одна неделя. Канал за каналом.",
                    mode="week",
                )

        build_legend_sheet(wb)

        fname    = (f"monthly_{ym}.xlsx" if report_type=="monthly"
                    else f"{report_type}_{ym}.xlsx")
        out_path = out_dir / fname
        wb.save(out_path)
        log.info(f"Файл сохранён: {out_path}")

        # Архивируем 24ч данные после месячного
        if report_type == "monthly":
            for ch in channels:
                archive_monthly(ch, ym)

        return [out_path]

    if tg_client is not None:
        files = await _execute(tg_client)
    else:
        _kw = get_telethon_kwargs()
        async with TelegramClient(SESSION_NAME, API_ID, API_HASH, **_kw) as _c:
            files = await _execute(_c)

    if not recipients or not tg_client:
        log.info("=== Готово ==="); return

    debug_tag = " [DEBUG]" if is_debug else ""
    caption   = f"📊{debug_tag} {report_type} | {period_label} | {len(files)} файл(ов)"
    for uid in recipients:
        try:
            await tg_client.send_message(uid, caption)
            for fp in files:
                await tg_client.send_file(uid, str(fp), caption=f"📎 {fp.name}")
                log.info(f"Отправлен {fp.name} → {uid}")
        except Exception as e:
            log.error(f"Ошибка отправки → {uid}: {e}")

    log.info("=== Готово ===")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--type",  required=True, choices=["daily","weekly","monthly"])
    parser.add_argument("--debug", action="store_true")
    args = parser.parse_args()
    asyncio.run(build_and_send(args.type, debug_override=args.debug))

if __name__ == "__main__":
    main()
